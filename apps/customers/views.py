from django.shortcuts import render
from django.contrib import messages
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.shortcuts import get_object_or_404, render, redirect
from django.views import View
from django.urls import reverse
from django.utils.html import escape
from django.http import JsonResponse
from django.db.models import Q
from apps.customers.models import PromotionCustomer, Promotions
from apps.customers.promotion_email import promotion_mail_send
from apps.property_management.models import PropertyManagementHotelRoom
from apps.users.models import Users
from solo_core.helpers.signer import URLEncryptionDecryption
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from datetime import datetime, timedelta
from django.http import HttpResponse
import openpyxl
import json
import pandas as pd
from io import BytesIO
import xlsxwriter
import io
from django.http import FileResponse
from openpyxl.styles import Font
import pytz
import os
from solo_core import settings
from apps.bookings.models import HotelRoomBookingCustomerDetails, RoomBookedCustomerDetails

#Downloading the customer details in excel format
class DownloadCustomerDataView(View):
    def get(self, request, *args, **kwargs):
        queryset = Users.objects.filter(Q(user_type='2') | Q(user_type='3'))

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Full Name', 'Email', 'Phone'])

        for item in queryset:
            worksheet.append([item.full_name, item.email, item.phone])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=customer_details.xlsx'
        workbook.save(response)

        return response
# Create your views here.
class CustomerManagementView(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {"breadcrumbs" : []}
        self.template = 'admin/home-page/customers-list/customer-listing.html' 
        self.context['title'] = 'Customers'
        self.generateBreadcrumbs()
        
    def get(self, request, *args, **kwargs):
        return render(request, self.template, context=self.context)

    def generateBreadcrumbs(self):
        self.context['breadcrumbs'].append({"name" : "Home", "route" : reverse('home:dashboard'),'active' : False})
        self.context['breadcrumbs'].append({"name" : "Customers", "route" : '','active' : True})
    
    # def post(self, request, *args, **kwargs):
    #     if 'download' in request.POST:
    #         queryset = Users.objects.filter(Q(user_type='2') | Q(user_type='3'))
    #         # Add any additional filters based on your requirements

    #         workbook = openpyxl.Workbook()
    #         worksheet = workbook.active
    #         worksheet.append(['Full Name', 'Email', 'Phone'])

    #         for item in queryset:
    #             worksheet.append([item.full_name, item.email, item.phone])

    #         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #         response['Content-Disposition'] = 'attachment; filename=customer_details.xlsx'
    #         workbook.save(response)

    #         return response

    #     # Handle regular DataTables request
    #     draw = request.POST.get('draw', 0)
    #     records_total = Users.objects.count()
    #     records_filtered = records_total  # You may need to adjust this based on filters
    #     data = [...]  # Your existing data preparation logic here

    #     response_data = {
    #         'draw': draw,
    #         'recordsTotal': records_total,
    #         'recordsFiltered': records_filtered,
    #         'data': data,
    #         'result': 'ok',
    #     }

    #     return HttpResponse(json.dumps(response_data), content_type='application/json')
        

class LoadCustomerManagementDatatable(BaseDatatableView):
    model = Users
    order_columns = ['id', 'image', 'full_name', 'email', 'phone', 'date_joined']

    def get_initial_queryset(self):
        filter_start_date = self.request.POST.get('columns[4][search][value]', None)
        filter_end_date = self.request.POST.get('columns[5][search][value]', None)

        queryset = self.model.objects.filter(user_type='2')


        if filter_start_date:
            start_date = datetime.strptime(filter_start_date, '%Y-%m-%d').date()
            queryset   = queryset.filter(Q(date_joined__gte=start_date))
 
        if filter_end_date:
            end_date = datetime.strptime(filter_end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(Q(date_joined__lte=end_date + timedelta(days=1)))

        filter_value = self.request.POST.get('columns[3][search][value]', None)
        if filter_value == '1':
            queryset = queryset.filter(is_active=True)
        elif filter_value == '2':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('-id')
    
    
    def filter_queryset(self, qs):
        search = self.request.POST.get('search[value]', None)
        if search:
            qs = qs.filter(Q(full_name__istartswith=search)|Q(email__istartswith=search)|Q(phone__istartswith=search)|Q(alternative_phone__istartswith=search))
        return qs

    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            phone_number = item.phone if item.phone else '------'
            alternative_phone = item.alternative_phone if item.alternative_phone else '------'
            json_data.append({
                'id'                  : escape(item.id),
                'encrypt_id'          : escape(URLEncryptionDecryption.enc(item.id)),
                'image'               : escape(self.request.build_absolute_uri(item.image.url)),
                'full_name'           : escape(item.full_name),
                'email'               : escape(item.email),
                'phone'               : escape(phone_number),
                'alternative_phone'   : escape(alternative_phone),
                'date_joined'         : escape(item.date_joined.strftime('%d-%m-%Y')),
                'is_active'           : escape(item.is_active),
            })
        return json_data


class CustomerManagementStatusChange(View):
    def __init__(self, **kwargs):
        self.response_format = {"status_code":101, "message":"", "error":""}
        
    def post(self, request, **kwargs):
        try:
            instance_id = request.POST.get('id', None)
            instance = Users.objects.get(id = instance_id)
            if instance_id:
                if instance.is_active:
                    instance.is_active = False
                else:
                    instance.is_active =True
                instance.save()
                self.response_format['status_code']=200
                self.response_format['message']= 'Success'
                
        except Exception as es:
            self.response_format['message']='error'
            self.response_format['error'] = str(es)
        return JsonResponse(self.response_format, status=200)
    



def generate_excel(request):
    
    try:
        data = json.loads(request.body)
        
        # Extract status, start_date, and end_date from the data
        status     = data.get('status')
        start_date = data.get('start_date')
        end_date   = data.get('end_date')
        instances  = Users.objects.filter(user_type='2')
        if status:
            if status == '1':
                instances = instances.filter(is_active=True)
            elif status == '2':
                instances = instances.filter(is_active=False)
            else:
                instances = instances.objects.all()
            
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            instances  = instances.filter(Q(date_joined__gte=start_date + timedelta(days=1)))
        
        if end_date:
            end_date  = datetime.strptime(end_date, '%Y-%m-%d').date()
            instances = instances.filter(Q(date_joined__lte=end_date + timedelta(days=1)))
        
        folder_name = 'customer_report'
        folder_path = os.path.join(settings.MEDIA_ROOT, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        excel_filename = 'Customer Reports.xlsx'
        excel_file_path = os.path.join(folder_path, excel_filename)
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Customers'
        
        # Write header row
        header = ['Email', 'Full Name', 'Phone Number', 'Date of Join']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)  # Apply bold font to header cells

        # Write data rows
        for row_num, instance in enumerate(instances, 2):
            row_data = [
                instance.email,
                instance.full_name,
                instance.phone,
                instance.date_joined.astimezone(pytz.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if instance.date_joined else None,
            ]
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Save the workbook to a temporary file
        workbook.save(excel_file_path)

        # Create a URL for the saved file
        excel_file_url = os.path.join(settings.MEDIA_URL, folder_name, excel_filename)

        # Return the URL in a JSON response
        return JsonResponse({'excel_file_url': excel_file_url})
    except Exception as es:
        return JsonResponse({'error': str(es)}, status=500)

    

# End Customer
    
# Promotion Start
    
class CustomerPromotionView(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {"breadcrumbs" : []}
        self.template = 'admin/home-page/customer-promotion/customer-promotion-list.html'
        self.context['title'] = 'Customer Promotion'
        self.generateBreadcrumbs()
        
    def get(self, request, *args, **kwargs):
        return render(request, self.template, context=self.context)

    def generateBreadcrumbs(self):
        self.context['breadcrumbs'].append({"name" : "Home", "route" : reverse('home:dashboard'),'active' : False})
        self.context['breadcrumbs'].append({"name" : "Customer Promotion", "route" : '','active' : True})
        

class LoadCustomerPromotionDatatable(BaseDatatableView):
    model = Promotions
    order_columns = ['id', 'title'] 
    
    def get_initial_queryset(self):
        filter_value = self.request.POST.get('columns[3][search][value]', None)
        if filter_value == '1':
            return self.model.objects.filter(is_active=True).order_by('-id')
        elif filter_value == '2':
            return self.model.objects.filter(is_active=False).order_by('-id')
        else:
            return Promotions.objects.all().order_by('-id')
    
    def filter_queryset(self, qs):
        search = self.request.POST.get('search[value]', None)
        if search:
            qs = qs.filter(Q(title__istartswith=search)|Q(subject__istartswith=search))
        return qs

    
    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            json_data.append({
                'id'              : escape(item.id),
                'title'            : escape(item.title),
                'subject'     : escape(item.subject),
                'message'     : escape(item.message),
                'encrypt_id'      : escape(URLEncryptionDecryption.enc(item.id)),
            })
        return json_data
    

class CustomerPromotionCreateOrUpdateView(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.context = {"breadcrumbs": []}
        self.context['title'] = 'Customer Promotion'
        self.action = "Created"
        self.template = 'admin/home-page/customer-promotion/create-or-update-customer-promotion.html'

    def get(self, request, *args, **kwargs):
        id = URLEncryptionDecryption.dec(kwargs.pop('id', None))
        user_queryset      = Users.objects.filter(Q(user_type='2') & Q(is_active=True))
        room_booking_obj   = RoomBookedCustomerDetails.objects.all()
        user_queryset_list = []
        email_set          = set() 

        for user in user_queryset:
            if user.email not in email_set:
                user_queryset_list.append({'user_id': user.id, 'full_name': user.full_name, 'email': user.email, 'table_type': '1'})
                email_set.add(user.email)

        for booking in room_booking_obj:
            if booking.email_address not in email_set:
                user_queryset_list.append({'user_id': booking.id, 'full_name': booking.full_name, 'email': booking.email_address, 'table_type': '2'})
                email_set.add(booking.email_address)
        
        if id:
            self.action = "Updated"
            self.context['promotion_obj'] = get_object_or_404(Promotions, id=id)
            self.context['promotion_customer_ids_email'] = PromotionCustomer.objects.filter(promotion_id=id).values_list('email', flat=True)

            
        self.context['user_queryset']     = user_queryset_list
        self.generateBreadcrumbs()
        return render(request, self.template, context=self.context)

    def generateBreadcrumbs(self):
        self.context['breadcrumbs'].append({"name": "Home", "route": reverse('home:dashboard'), 'active': False})
        self.context['breadcrumbs'].append(
            {"name": "Customer Promotion", "route": reverse('customer_management:customer_promotion.index'), 'active': False})
        self.context['breadcrumbs'].append({"name": "{} Customer Promotion".format(self.action), "route": '', 'active': True})

    def post(self, request, *args, **kwargs):
        customer_promotion_id = request.POST.get('customer_promotion_id', None)
        try:
            if customer_promotion_id:
                self.action = 'Updated'
                promotion_obj         = get_object_or_404(Promotions, id=customer_promotion_id)
            else:
                promotion_obj         = Promotions()
            promotion_obj.title       = request.POST.get('title')
            promotion_obj.subject     = request.POST.get('subject')
            promotion_obj.message     = request.POST.get('message')
            customer_list             = request.POST.getlist('customers')
            
            if 'all' in customer_list:
                user_queryset      = Users.objects.filter(Q(user_type='2') & Q(is_active=True))
                room_booking_obj   = RoomBookedCustomerDetails.objects.all()
                user_queryset_list = []
                email_set          = set() 
                for user in user_queryset:
                    if user.email not in email_set:
                        user_queryset_list.append({'user_id': user.id, 'full_name': user.full_name, 'email': user.email, 'table_type': '1'})
                        email_set.add(user.email)

                for booking in room_booking_obj:
                    if booking.email_address not in email_set:
                        user_queryset_list.append({'user_id': booking.id, 'full_name': booking.full_name, 'email': booking.email_address, 'table_type': '2'})
                        email_set.add(booking.email_address)
            else:
                user_queryset_list              = request.POST.getlist('customers')
                user_queryset_list                       = [json.loads(item.replace("'", "\"")) for item in user_queryset_list]
                
                
                
            if request.FILES.__len__() != 0:
                if request.POST.get('promotion_image', None) is None:
                    promotion_obj.image = request.FILES.get('promotion_image')
            promotion_obj.save()
            

            email_values = []
            email_list   = []
            for data_dict in user_queryset_list:
                email_values.append(data_dict.get('email'))
                promotion_cust_obj = PromotionCustomer(promotion=promotion_obj, email=data_dict.get('email'), full_name=data_dict.get('full_name'))
                
                if data_dict.get('table_type') == '1':
                    promotion_cust_obj.table_type = '1'
                else:
                    promotion_cust_obj.table_type = '2'
                email_list.append([data_dict.get('email'), data_dict.get('full_name')])
                promotion_cust_obj.save()
            if not customer_promotion_id:
                promotion_mail_send(request, promotion_obj, email_list)

            messages.success(request, f"Data Successfully "+ self.action)

        except Exception as e:
            messages.error(request, f"Something went wrong."+str(e))
            if customer_promotion_id is not None:
                return redirect('customer_management:customer_promotion.update', id=URLEncryptionDecryption.dec(int(customer_promotion_id)))
            return redirect('customer_management:customer_promotion.create')
        return redirect('customer_management:customer_promotion.index')
    

class DestroyCustomerPromotionRecordsView(View):
    def __init__(self, **kwargs):
        self.response_format = {"status_code": 101, "message": "", "error": ""}

    def post(self, request, *args, **kwargs):
        try:
            property_collection_id = request.POST.getlist('ids[]')
            if property_collection_id:
                Promotions.objects.filter(id__in=property_collection_id).delete()
                self.response_format['status_code'] = 200
                self.response_format['message'] = 'Success'
        except Exception as e:
            self.response_format['message'] = 'error'
            self.response_format['error'] = str(e)
            
        return JsonResponse(self.response_format, status=200)
        

# End

# Start Room Booked customers

class RoomBookedCustomerManagementView(View):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.context = {"breadcrumbs" : []}
        self.template = 'admin/home-page/booked-customer-list/booked-customer-listing.html' 
        self.context['title'] = 'Customers'
        self.generateBreadcrumbs()
        
    def get(self, request, *args, **kwargs):
        return render(request, self.template, context=self.context)

    def generateBreadcrumbs(self):
        self.context['breadcrumbs'].append({"name" : "Home", "route" : reverse('home:dashboard'),'active' : False})
        self.context['breadcrumbs'].append({"name" : "Customers", "route" : '','active' : True})
    


class LoadRoomBookedCustomerManagementDatatable(BaseDatatableView):
    model         = HotelRoomBookingCustomerDetails
    order_columns = ['id']

    def get_initial_queryset(self):
        filter_start_date = self.request.POST.get('columns[4][search][value]', None)
        filter_end_date   = self.request.POST.get('columns[5][search][value]', None)

        queryset = self.model.objects.all()

        if filter_start_date:
            start_date = datetime.strptime(filter_start_date, '%Y-%m-%d').date()
            queryset   = queryset.filter(Q(booked_room__check_in__gte=start_date + timedelta(days=1)))
        

        if filter_end_date:
            end_date   = datetime.strptime(filter_end_date, '%Y-%m-%d').date()
            queryset   = queryset.filter(Q(booked_room__check_out__lte=end_date + timedelta(days=1)))

            
        return queryset.order_by('-id')
    
    def filter_queryset(self, qs):
        search = self.request.POST.get('search[value]', None)
        if search:
            qs = qs.filter(Q(guest_details__full_name__istartswith=search)|Q(guest_details__email__istartswith=search))
        return qs
    
    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            property_obj              = PropertyManagementHotelRoom.objects.filter(hotel_room=item.booked_room.room_id).first()
            json_data.append({
                'id'                  : escape(item.id),
                'encrypt_id'          : escape(URLEncryptionDecryption.enc(item.id)),
                'property_name'       : property_obj.property_management.name if property_obj else '',
                'roll_number'         : escape(item.booked_room.room_sl_no),
                'name'                : escape(item.guest_details.full_name),
                'email'               : escape(item.guest_details.email_address),
                'phone'               : escape(item.guest_details.mobile_number),
                'check_in'            : escape(item.booked_room.check_in.strftime('%d-%m-%Y')),
                'check_out'           : escape(item.booked_room.check_out.strftime('%d-%m-%Y')),
                'document'            : escape(self.request.build_absolute_uri(item.guest_details.document.url)),
            })
        return json_data
    
def generate_room_booked_excel(request):
    
    try:
        data = json.loads(request.body)
        # Extract status, start_date, and end_date from the data
        start_date = data.get('start_date')
        end_date   = data.get('end_date')
        instances  = HotelRoomBookingCustomerDetails.objects.all().order_by('-id')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            instances   = instances.filter(Q(booked_room__check_in__gte=start_date + timedelta(days=1)))
        

        if end_date:
            end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
            instances   = instances.filter(Q(booked_room__check_out__lte=end_date + timedelta(days=1)))
        
        folder_name = 'Booked Customer Report'
        folder_path = os.path.join(settings.MEDIA_ROOT, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        excel_filename = 'bookedcustomersreport.xlsx'
        excel_file_path = os.path.join(folder_path, excel_filename)
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Booked Customers'
        

        header = ['Property Name', 'Room Number', 'Full Name', 'Email', 'Phone Number', 'Check In', 'Check Out']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)  # Apply bold font to header cells

        # Write data rows
        for row_num, instance in enumerate(instances, 2):
            property_obj = get_object_or_404(PropertyManagementHotelRoom, hotel_room=instance.booked_room.room_id)
            row_data = [
                property_obj.property_management.name if property_obj else '',
                instance.booked_room.room_sl_no,
                instance.guest_details.full_name,
                instance.guest_details.email_address,
                instance.guest_details.mobile_number,
                instance.booked_room.check_in.astimezone(pytz.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if instance.booked_room.check_in else None,
                instance.booked_room.check_out.astimezone(pytz.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if instance.booked_room.check_out else None,
                
            ]
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Save the workbook to a temporary file
        workbook.save(excel_file_path)

        # Create a URL for the saved file
        excel_file_url = os.path.join(settings.MEDIA_URL, folder_name, excel_filename)

        # Return the URL in a JSON response
        return JsonResponse({'excel_file_url': excel_file_url})
    except Exception as es:
        return JsonResponse({'error': str(es)}, status=500)
