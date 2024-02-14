from apps.customers.api.schemas import GetBookedCustomerDetailsExcelDownloadSchemas
from apps.customers.api.serializers import BookedCustomerDetailsExcelDownloadSerializer
from apps.bookings.models import RoomBookedCustomerDetails
from drf_yasg import openapi
from rest_framework.response import Response
from drf_yasg.utils import swagger_auto_schema
from solo_core.helpers.pagination import RestPagination
import django_filters
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import generics, filters
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

# Create your views here.
# class CustomerDetailsFilter(django_filters.FilterSet):
#     created_date__gte = django_filters.DateTimeFilter(field_name='created_date', lookup_expr='gte')
#     created_date__lte = django_filters.DateTimeFilter(field_name='created_date', lookup_expr='lte')

#     class Meta:
#         model = RoomBookedCustomerDetails
#         fields = []

class GetBookedCustomerDetailsExcelDownloadAPIView(generics.ListAPIView):
    queryset = RoomBookedCustomerDetails.objects.all().order_by('-id')
    serializer_class = BookedCustomerDetailsExcelDownloadSerializer
    pagination_class = RestPagination
    filter_backends = [filters.SearchFilter, django_filters.rest_framework.DjangoFilterBackend]
    search_fields = ['name']
    # filterset_class = CustomerDetailsFilter

    @swagger_auto_schema(tags=["CustomerDetailsExcelDownload"])
    def post(self, request):
        from_date   = request.data.get('from_date', None)
        to_date     = request.data.get('to_date', None)

        if not from_date or not to_date:
            return Response({"detail": "Both from_date and to_date are required."}, status=400)

        # Convert from_date and to_date to datetime objects
        try:
            from_date   = datetime.strptime(from_date, '%Y-%m-%d')
            to_date     = datetime.strptime(to_date, '%Y-%m-%d')
        except ValueError:
            return Response({"detail": "Invalid date format. Please use 'YYYY-MM-DD'."}, status=400)

        # Ensure the end date is set to the end of the day
        to_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)

        queryset = self.get_queryset()

        # Filter by created_date within the date range
        queryset = queryset.filter(Q(created_date__gte=from_date) & Q(created_date__lte=to_date))

        search_term = request.data.get('search')
        if search_term:
            queryset = queryset.filter(name__icontains=search_term)

        page = self.paginate_queryset(queryset)
        serializer = GetBookedCustomerDetailsExcelDownloadSchemas(page, many=True, context={"request": request})

        # Generate Excel file and return as a response
        excel_file_response = generate_excel_file(serializer.data, from_date, to_date)
        return excel_file_response

def generate_excel_file(data, from_date, to_date):
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add column headers including Serial Number
    headers = ['Serial No', 'Full Name', 'Email Address', 'Mobile Number', 'Date of Birth', 'Guest Address']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Populate the worksheet with data and serial numbers
    for row_num, record in enumerate(data, 2):
        ws[f"A{row_num}"] = row_num - 1  # Serial number starts from 1
        ws[f"B{row_num}"] = record.get('full_name', '')
        ws[f"C{row_num}"] = record.get('email_address', '')
        ws[f"D{row_num}"] = record.get('mobile_number', '')
        ws[f"E{row_num}"] = record.get('dob', '')
        ws[f"F{row_num}"] = record.get('guest_address', '')

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Customer_details_{from_date.strftime("%d %m %Y")}_to_{to_date.strftime("%d %m %Y")}.xlsx'
    wb.save(response)

    return response
